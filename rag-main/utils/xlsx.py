import os
import shutil
from zipfile import ZipFile, BadZipFile
import openpyxl
from PIL import Image
import olefile
from oletools import oleobj
import subprocess

class GetOleFileUtil:
    def __init__(self):
        pass

    def save_olefile(self, file, savepath):
        name_dict = {
            'xlsx': 'xl',
            'docx': 'word',
            'pptx': 'ppt'
        }

        try:
            # 检查文件后缀
            ext = os.path.splitext(file)[-1].replace('.', '')
            name = name_dict.get(ext, '')
            if not name:
                return False, '文件格式不受支持'

            # 创建保存路径
            savefolder = self.create_folder(file, savepath)

            # 确保这是一个可解压的ZIP文件
            try:
                with ZipFile(file, 'r') as zip:
                    embedded_files = []
                    for entry in zip.infolist():
                        if entry.filename.startswith(f'{name}/embeddings/'):
                            if entry.filename.endswith('.bin') and 'ole' in entry.filename.lower():
                                embedded_files.append(entry.filename)

                    if not embedded_files:
                        return False, '未找到 OLE 对象'

                    print(f'发现 OLE 对象文件：{embedded_files}')

                    for entry_filename in embedded_files:
                        with zip.open(entry_filename) as f:
                            if not olefile.isOleFile(f):
                                continue
                            with olefile.OleFileIO(f) as ole:
                                stream = None
                                try:
                                    stream = ole.openstream('\x01Ole10Native')
                                    opkg = oleobj.OleNativeStream(stream)
                                except IOError:
                                    print('不是ole文件')
                                    if stream is not None:
                                        stream.close()
                                    continue
                                if opkg.is_link:
                                    print('是链接不是文件，跳过')
                                    continue
                                ole_filename = self.re_decode(opkg.filename)
                                filename = os.path.join(savefolder, ole_filename)
                                try:
                                    print(f'导出ole中的文件：{filename}')
                                    with open(filename, 'wb') as writer:
                                        n_dumped = 0
                                        next_size = min(oleobj.DUMP_CHUNK_SIZE, opkg.actual_size)
                                        while next_size:
                                            data = stream.read(next_size)
                                            writer.write(data)
                                            n_dumped += len(data)
                                            if len(data) != next_size:
                                                break
                                            next_size = min(oleobj.DUMP_CHUNK_SIZE, opkg.actual_size - n_dumped)
                                except Exception as exc:
                                    print('在转存时出现错误：')
                                    raise exc
                                finally:
                                    if stream is not None:
                                        stream.close()
                return True, "提取成功"
            except BadZipFile:
                return False, '文件不是有效的zip格式'
        except Exception as ex:
            return False, f'提取ole文件异常：{ex}'

    @staticmethod
    def re_decode(s, encoding='gbk'):
        i81 = s.encode('iso-8859-1')
        return i81.decode(encoding)

    @staticmethod
    def create_folder(file, savepath):
        filename = os.path.basename(os.path.splitext(file)[0])
        new_folder = os.path.join(savepath, filename)
        if not os.path.exists(new_folder):
            os.makedirs(new_folder)
            print(f'创建文件夹成功：{new_folder}')
        else:
            print(f'文件夹已存在，删除再创建：{new_folder}')
            shutil.rmtree(new_folder)
            os.makedirs(new_folder)
        return new_folder

def print_sheet_content(sheet):
    """打印工作表的内容，并将其保存到数据列表中"""
    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        if all(value is None for value in row):
            continue  # 跳过全为空的行
        row_data = [str(value) if value is not None else '' for value in row]
        sheet_data.append(row_data)
    return sheet_data


def convert_emf_to_png(emf_path, png_path):
    """使用 ImageMagick 将 EMF 转换为 PNG"""
    subprocess.run(["magick", "convert", emf_path, png_path])

def extract_images_from_excel(file_path, output_dir):
    """从Excel文件中提取嵌入的图片并保存到指定目录，排除 OLE 对象"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    image_found = False  # 用于标记是否找到图片

    with ZipFile(file_path, 'r') as zip:
        for entry in zip.infolist():
            # 只处理 xl/media/ 目录中的文件，排除包含 ole 关键字的文件
            if entry.filename.startswith('xl/media/') and 'ole' not in entry.filename.lower():
                image_found = True

                # 获取文件名并检查其有效性
                file_name = os.path.basename(entry.filename)
                if not file_name:  # 如果文件名为空，跳过处理
                    continue

                with zip.open(entry.filename) as image_file:
                    try:
                        # 保存原始文件到临时路径
                        temp_image_path = os.path.join(output_dir, file_name)
                        
                        with open(temp_image_path, 'wb') as f:
                            f.write(image_file.read())

                        # 尝试用Pillow处理
                        try:
                            image = Image.open(temp_image_path)
                            image.save(temp_image_path.replace('.emf', '.png'))  # 保存为 PNG
                            print(f"图片保存到：{temp_image_path.replace('.emf', '.png')}")
                        except IOError:
                            # 如果是 .emf 文件，使用 ImageMagick 转换
                            if entry.filename.lower().endswith('.emf'):
                                convert_emf_to_png(temp_image_path, temp_image_path.replace('.emf', '.png'))
                                print(f".emf 文件转换为 PNG 并保存到：{temp_image_path.replace('.emf', '.png')}")
                            else:
                                print(f"无法处理图片 {entry.filename}: {e}")
                    except Exception as e:
                        print(f"无法处理图片 {entry.filename}: {e}")

    if not image_found:
        print("没有在文件中找到任何图片")



def process_excel_files(file_path):
    """处理Excel文件，并返回所需数据"""
    output_dir = os.path.join(os.path.dirname(file_path), "output")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    excel_data = []

    ole_util = GetOleFileUtil()

    # 提取OLE对象
    ole_result = ole_util.save_olefile(file_path, output_dir)
    if not ole_result[0]:
        print(f"OLE 提取失败：{ole_result[1]}")

    try:
        # 获取工作簿对象
        workbook = openpyxl.load_workbook(file_path)
        sheetnames = workbook.sheetnames

        for sheetname in sheetnames:
            worksheet = workbook[sheetname]
            sheet_content = print_sheet_content(worksheet)

            # 提取图片
            extract_images_from_excel(file_path, output_dir)

            # 保存提取的信息
            sheet_data = {
                'file': os.path.basename(file_path),
                'sheet': sheetname,
                'content': sheet_content,
                'images': output_dir  # 图片路径
            }
            excel_data.append(sheet_data)

    except FileNotFoundError as fnf_error:
        print(fnf_error)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("文件无效，请提供一个有效的 Excel 文件。")
    except Exception as e:
        print(f"发生错误：{e}")

    return excel_data

